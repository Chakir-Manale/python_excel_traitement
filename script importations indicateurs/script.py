#!/usr/bin/env python
# coding: utf-8

# In[2]:


import pandas as pd
import openpyxl as xl
import numpy as np
from openpyxl import load_workbook

countries_file = open("inputs\\countries_code.txt","r")
countries = countries_file.readlines() 

cible_file_name = "Matrice_2ème problématique.xlsx"
wb = load_workbook(filename = cible_file_name)
sheet = wb.active 

source_file = open("inputs\\liste_indicateurs.txt","r")
ligne_indicateurs = source_file.readlines() 

column_index = 4;
for ligne in ligne_indicateurs :
    indicateur = ligne.split(":")
    print(indicateur[1] +"en cours ..")
    source_file_name ="indicateurs\\"+ indicateur[0]+".xls"
    excel_data = pd.read_excel(source_file_name, sheet_name='Data',skiprows=3)
    country_start = 4
    for country in countries:
        df_country = excel_data[excel_data["Country Code"] == country.replace('\n','')]
        df_country.replace(np.nan, 0)
        index_annee = 34
        sheet.cell(row = 3 , column = column_index).value = indicateur[1]
        for row in range(country_start, country_start+29 ):
            sheet.cell(row = row , column = column_index).value = df_country.iat[0, index_annee]
            index_annee+=1
        country_start+=29
    column_index+=1

wb.save("Matrice_2ème problématique.xlsx")
print("\n------------ le processus est ternimé -------------\n")
countries_file.close()
source_file.close() 
input("appuyez sur une touche pour quitter")

